import openpyxl

wk = openpyxl.load_workbook('C:\\Users\Laptop\Documents\Testdata1.xlsx')

def fetch_number_rows(Sheetname):
    sh = wk[Sheetname]
    return sh.max_row
def fetch_cell_data(Sheetname, row, cell):
    sh = wk[Sheetname]
    cell = sh.cell(int(row), int(cell))
    return cell.value
print(fetch_number_rows("Blad1"))
print(fetch_cell_data('Blad1', 1, 1))
